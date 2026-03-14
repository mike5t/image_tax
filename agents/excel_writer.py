"""Agent 7: Excel Append Writer.

Manages the Transactions.xlsx workbook — creates sheets on first run,
always appends (never overwrites), and routes rows to category-specific sheets.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import config
from models.transaction import Transaction

logger = logging.getLogger(__name__)

# Column headers — each line item gets its own row
HEADERS = [
    "Date",
    "Merchant",
    "Currency",
    "Item Description",
    "Qty",
    "Unit Price",
    "Item Amount",
    "Receipt Total",
    "VAT",
    "Category",
    "Receipt #",
    "Payment Method",
    "Business Use",
    "Confidence",
    "Image Path",
    "Text Path",
    "Telegram File ID",
    "Message ID",
    "Recorded At",
]

# All sheets that should exist
_ALL_SHEETS = [
    "Transactions",
    *config.CATEGORY_SHEET_MAP.values(),
    "ForeignCurrency",
    "Rejected",
]


def _ensure_workbook(path: Path) -> Workbook:
    """Open existing workbook or create a new one with all required sheets."""
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Remove the default "Sheet"
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Ensure all sheets exist with headers
    for sheet_name in _ALL_SHEETS:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(HEADERS)
            logger.info("Created sheet: %s", sheet_name)

    return wb


def _txn_to_rows(txn: Transaction) -> list[list]:
    """Convert a Transaction to one or more rows — one per line item.

    If there are no items, a single summary row is written.
    """
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = [
        txn.date,
        txn.merchant,
        txn.currency,
    ]
    tail = [
        txn.total,
        txn.vat if txn.vat is not None else "",
        txn.category,
        txn.receipt_number,
        txn.payment_method,
        txn.business_use,
        round(txn.confidence, 2),
        txn.image_path,
        txn.raw_text_path,
        txn.telegram_file_id,
        txn.telegram_message_id,
        now_str,
    ]

    rows: list[list] = []
    if txn.items:
        for item in txn.items:
            rows.append(base + [
                item.description,
                item.quantity,
                item.unit_price,
                item.total,
            ] + tail)
    else:
        # No items — single row with empty item columns
        rows.append(base + ["(whole receipt)", 1, txn.total, txn.total] + tail)

    return rows


def append_transaction(txn: Transaction, rejected: bool = False) -> str:
    """Append transaction rows to the master sheet + category sheet.

    Each line item becomes its own row so every amount is visible.

    Args:
        txn: validated, confirmed Transaction
        rejected: if True, write to Rejected sheet instead of category sheet

    Returns:
        Summary string for logging / user feedback.
    """
    wb = _ensure_workbook(config.EXCEL_PATH)
    rows = _txn_to_rows(txn)
    sheets_written: list[str] = []

    if rejected:
        for row in rows:
            wb["Rejected"].append(row)
        sheets_written.append("Rejected")
    else:
        # Always write to master
        for row in rows:
            wb["Transactions"].append(row)
        sheets_written.append("Transactions")

        # Category-specific sheet
        cat_sheet = config.CATEGORY_SHEET_MAP.get(txn.category)
        if cat_sheet and cat_sheet in wb.sheetnames:
            for row in rows:
                wb[cat_sheet].append(row)
            sheets_written.append(cat_sheet)

        # Foreign currency sheet
        if txn.foreign_currency and "ForeignCurrency" in wb.sheetnames:
            for row in rows:
                wb["ForeignCurrency"].append(row)
            sheets_written.append("ForeignCurrency")

    wb.save(config.EXCEL_PATH)
    item_count = len(txn.items) if txn.items else 1
    summary = f"Written {item_count} item(s) to: {', '.join(sheets_written)}"
    logger.info(summary)
    return summary


def test_write():
    """Quick test — creates a dummy row and saves."""
    txn = Transaction(
        date="2026-02-17",
        merchant="Test Merchant",
        currency="ZAR",
        total=99.99,
        category="Groceries",
        confidence=0.95,
        needs_review=False,
        image_path="test.jpg",
        raw_text_path="test.txt",
    )
    result = append_transaction(txn)
    print(f"✅ Test write complete. {result}")
    print(f"   File: {config.EXCEL_PATH}")
