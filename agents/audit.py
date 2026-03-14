"""Agent 8: Audit Assistant.

Provides duplicate detection, missing evidence checks, anomaly detection,
and monthly summaries. Can be run via Telegram /audit command or scheduled.
"""

from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

import config
from agents.excel_writer import HEADERS

logger = logging.getLogger(__name__)


def _load_transactions() -> list[dict]:
    """Load all rows from the Transactions sheet as dicts."""
    if not config.EXCEL_PATH.exists():
        return []

    wb = load_workbook(config.EXCEL_PATH, read_only=True)
    if "Transactions" not in wb.sheetnames:
        return []

    ws = wb["Transactions"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    result = []
    for row in rows:
        if row and len(row) >= len(HEADERS):
            d = dict(zip(HEADERS, row))
            result.append(d)
    return result


def detect_duplicates() -> list[str]:
    """Find rows with same merchant + total + date within 24 hours."""
    txns = _load_transactions()
    warnings: list[str] = []

    # Group by (merchant, total)
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for t in txns:
        key = (str(t.get("Merchant", "")).lower(), t.get("Total", 0))
        groups[key].append(t)

    for key, group in groups.items():
        if len(group) > 1:
            # Check if dates are within 24h
            dates = []
            for t in group:
                try:
                    dates.append(datetime.strptime(str(t.get("Date", "")), "%Y-%m-%d"))
                except ValueError:
                    pass
            if len(dates) >= 2:
                dates.sort()
                for i in range(1, len(dates)):
                    if dates[i] - dates[i - 1] <= timedelta(hours=24):
                        warnings.append(
                            f"⚠️ Possible duplicate: {key[0]} — {key[1]} on "
                            f"{dates[i-1].date()} and {dates[i].date()}"
                        )
    return warnings


def detect_missing_evidence() -> list[str]:
    """Find rows where the image file doesn't exist on disk."""
    txns = _load_transactions()
    warnings: list[str] = []

    for t in txns:
        img = t.get("Image Path", "")
        if img and not Path(img).exists():
            warnings.append(f"📁 Missing image: {img} (merchant: {t.get('Merchant', '?')})")
    return warnings


def detect_anomalies() -> list[str]:
    """Flag unusual totals or categories."""
    txns = _load_transactions()
    warnings: list[str] = []

    for t in txns:
        total = t.get("Total", 0)
        if isinstance(total, (int, float)):
            if total > 10_000:
                warnings.append(
                    f"💰 High amount: {t.get('Merchant', '?')} — "
                    f"{t.get('Currency', 'ZAR')} {total:.2f}"
                )
            if total < 0:
                warnings.append(
                    f"🔴 Negative amount: {t.get('Merchant', '?')} — {total:.2f}"
                )
        if t.get("Category") == "Unknown":
            warnings.append(f"❓ Uncategorized: {t.get('Merchant', '?')} on {t.get('Date', '?')}")

    return warnings


def monthly_summary(year: int | None = None, month: int | None = None) -> str:
    """Generate a summary of totals by category for a given month."""
    now = datetime.now()
    year = year or now.year
    month = month or now.month
    month_prefix = f"{year}-{month:02d}"

    txns = _load_transactions()
    totals: dict[str, float] = defaultdict(float)
    count = 0

    for t in txns:
        if str(t.get("Date", "")).startswith(month_prefix):
            cat = str(t.get("Category", "Unknown"))
            amt = t.get("Total", 0)
            if isinstance(amt, (int, float)):
                totals[cat] += amt
                count += 1

    if not totals:
        return f"📊 No transactions found for {month_prefix}."

    lines = [f"📊 *Monthly Summary — {month_prefix}*", f"Transactions: {count}", ""]
    grand = 0.0
    for cat in sorted(totals.keys()):
        lines.append(f"  • {cat}: R {totals[cat]:,.2f}")
        grand += totals[cat]
    lines.append(f"\n  *Grand Total: R {grand:,.2f}*")

    return "\n".join(lines)


def run_full_audit() -> str:
    """Run all audit checks and return a combined report."""
    sections: list[str] = ["🔍 *Audit Report*\n"]

    dupes = detect_duplicates()
    if dupes:
        sections.append("*Possible Duplicates:*")
        sections.extend(dupes)
    else:
        sections.append("✅ No duplicates found.")

    missing = detect_missing_evidence()
    if missing:
        sections.append("\n*Missing Evidence:*")
        sections.extend(missing)
    else:
        sections.append("✅ All evidence files present.")

    anomalies = detect_anomalies()
    if anomalies:
        sections.append("\n*Anomalies:*")
        sections.extend(anomalies)
    else:
        sections.append("✅ No anomalies detected.")

    sections.append("\n" + monthly_summary())

    return "\n".join(sections)
